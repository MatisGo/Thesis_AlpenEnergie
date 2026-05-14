"""
add_swissgrid_prices.py
-----------------------
One-time script: merges Swissgrid balance energy prices into matis_2025_.xlsx
as two new columns (BG_Long_EUR_MWh, BG_Short_EUR_MWh) at columns T and U.

Source : swissgrid-balance-energy-prices-2025-full (1).xlsx  (15-min, ct/kWh)
Target : matis_2025_.xlsx  (5-min, adds EUR/MWh columns)

Unit conversion : ct/kWh × 10 = EUR/MWh

Run once:
  python add_swissgrid_prices.py
"""

import os
import pandas as pd
from openpyxl import load_workbook

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
MATIS_PATH = os.path.join(SCRIPT_DIR, 'Input', 'matis_2025_.xlsx')
SG_PATH    = os.path.join(SCRIPT_DIR, 'Input', 'swissgrid-balance-energy-prices-2025-full (1).xlsx')

NEW_COL_LONG  = 20   # Excel column T (1-indexed)
NEW_COL_SHORT = 21   # Excel column U (1-indexed)


def main():
    # 1. Load Swissgrid data and convert to EUR/MWh (ct/kWh × 10)
    print('Loading Swissgrid balance energy prices ...')
    df_sg = pd.read_excel(SG_PATH)
    df_sg['Timestamp'] = pd.to_datetime(df_sg['Timestamp'])
    df_sg = df_sg.drop_duplicates('Timestamp').sort_values('Timestamp').set_index('Timestamp')
    df_sg['BG_Long_EUR_MWh']  = (df_sg['BG-long (ct/kWh)']  * 10.0).round(4)
    df_sg['BG_Short_EUR_MWh'] = (df_sg['BG-short (ct/kWh)'] * 10.0).round(4)
    print(f'  {len(df_sg):,} rows  |  '
          f'{df_sg.index.min().date()} → {df_sg.index.max().date()}')

    # 2. Resample from 15-min to 5-min by forward-fill
    #    (a 15-min price applies uniformly to all three 5-min slots it covers)
    idx_5min = pd.date_range(
        start=df_sg.index.min(),
        end=df_sg.index.max() + pd.Timedelta('14min'),
        freq='5min')
    df_5min = df_sg[['BG_Long_EUR_MWh', 'BG_Short_EUR_MWh']].reindex(
        idx_5min, method='ffill')

    # Build lookup: "dd.mm.yyyy HH:MM:SS" → (long, short)
    price_map = {
        ts.strftime('%d.%m.%Y %H:%M:%S'): (row['BG_Long_EUR_MWh'], row['BG_Short_EUR_MWh'])
        for ts, row in df_5min.iterrows()
    }
    print(f'  5-min lookup entries: {len(price_map):,}')

    # 3. Open Excel and write new columns
    print('Opening matis_2025_.xlsx ...')
    wb = load_workbook(MATIS_PATH)
    ws = wb.active

    if ws.cell(row=1, column=NEW_COL_LONG).value is not None:
        print(f'  NOTE: column {NEW_COL_LONG} already has data — overwriting.')

    # Row 1: labels
    ws.cell(row=1, column=NEW_COL_LONG).value  = 'BG Long Price'
    ws.cell(row=1, column=NEW_COL_SHORT).value = 'BG Short Price'
    # Row 2: units
    ws.cell(row=2, column=NEW_COL_LONG).value  = 'EUR/MWh'
    ws.cell(row=2, column=NEW_COL_SHORT).value = 'EUR/MWh'
    # Row 3: internal name (used by hydro_constants.py)
    ws.cell(row=3, column=NEW_COL_LONG).value  = 'BG_Long_EUR_MWh'
    ws.cell(row=3, column=NEW_COL_SHORT).value = 'BG_Short_EUR_MWh'

    # Data rows (row 4 onwards)
    print('Writing price data to rows ...')
    written = missing = 0
    for r in range(4, ws.max_row + 1):
        ts_raw = ws.cell(row=r, column=1).value
        if ts_raw is None:
            continue
        ts_str = str(ts_raw).strip()
        if ts_str in price_map:
            long_p, short_p = price_map[ts_str]
            ws.cell(row=r, column=NEW_COL_LONG).value  = long_p
            ws.cell(row=r, column=NEW_COL_SHORT).value = short_p
            written += 1
        else:
            missing += 1

    print(f'  Written : {written:,} rows')
    if missing:
        print(f'  Missing : {missing:,} rows (timestamps outside 2025 Swissgrid range)')

    wb.save(MATIS_PATH)
    print(f'Saved : {MATIS_PATH}')


if __name__ == '__main__':
    main()