"""
get_load_data.py
----------------
Two jobs in one run:

  1. EXTEND CALENDAR ROWS
     Appends new 5-min rows to Data_Prediction.xlsx up to TODAY + EXTEND_DAYS_AHEAD.
     Fills the first 5 calendar columns only (Time, Date, Day_Time, Weekday, PHolyday).
     Sensor columns are left empty for the SCADA import (Job 2).

  2. IMPORT SCADA CSV DATA
     Reads every *.csv file in the Input/ folder.
     CSV format (semicolon-delimited):
       Row 1  — units  (skipped)
       Row 2  — signal names  (matched against Excel row 3 to find target columns)
       Row 3+ — data  (timestamp dd.mm.yyyy HH:MM:SS ; values…)
     For each data row the matching Excel row is found by the Time string and the
     signal values are written to the correct columns.
     NOTE: 00:00:00 timestamps are treated as the START of that calendar day
           (matching the convention already used in Data_Prediction.xlsx).
     Existing Weekday formulas (=WEEKDAY…) are never overwritten.
"""

import os
import glob
import datetime
import pandas as pd
from openpyxl import load_workbook

# ── Config ────────────────────────────────────────────────────────────────────
SCRIPT_DIR        = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH        = os.path.join(SCRIPT_DIR, 'Ressource', 'Data_Prediction.xlsx')
HOLIDAY_CSV       = os.path.join(SCRIPT_DIR, 'Ressource', 'Public_Holiday_Meiringen_2025.csv')
INPUT_DIR         = os.path.join(SCRIPT_DIR, '..', 'Input')
SHEET_NAME        = 'Data_January'
EXTEND_DAYS_AHEAD = 10    # generate calendar rows up to today + this many days
STEP_MINUTES      = 5     # resolution of the Excel (5-min)
SIGNALNAME_ROW    = 3     # openpyxl row number containing signal names in the Excel
FIRST_DATA_ROW    = 4     # openpyxl row number of first actual data row
GAP_CHECK_DAYS    = 14    # how many past days to check for missing data

# Canonical Excel name for the net load signal (used by the gap check)
LOAD_IS_SIGNAL = 'PR_MI_SDR___:Prog:EWO_Last:IST_wert/t'

# Signal aliases: maps current SCADA export names → canonical Excel row-3 names.
# Add entries here whenever the SCADA system renames a signal.
SIGNAL_ALIASES = {
    # Renamed 2026 — Load_Is (net consumption)
    'PR_MI_SDR___:Prog:EWO:Last_Netto:IST:IST_wert/t':  'PR_MI_SDR___:Prog:EWO_Last:IST_wert/t',
    # Renamed 2026 — Production forecast (col F)
    'PR_MI_SDR___:IO:Slot2:Prog_EWO_EGSHydro:IST_wert/t': 'PR_MI_SDR___:IO:Slot2:Prog_EWO_Prod:IST_wert/t',
    # Renamed 2026 — Load forecast (col G)
    'PR_MI_SDR___:IO:Slot2:Prog_EWO_LGS:IST_wert/t':    'PR_MI_SDR___:IO:Slot2:Prog_EWO_Last:IST_wert/t',
}
# ─────────────────────────────────────────────────────────────────────────────


# ── Helper: normalise a raw timestamp cell value to "dd.mm.yyyy HH:MM:SS" ───
def _norm_ts(raw) -> str | None:
    """Convert any Excel cell value to the canonical timestamp string used as key."""
    if raw is None:
        return None
    if isinstance(raw, datetime.datetime):
        return raw.strftime('%d.%m.%Y %H:%M:%S')
    s = str(raw).strip()
    if not s or s in ('Time ', 'Time', 'Einheit', 'Signalname'):
        return None
    # Already in the right format dd.mm.yyyy HH:MM:SS
    try:
        datetime.datetime.strptime(s, '%d.%m.%Y %H:%M:%S')
        return s
    except ValueError:
        return None


# ── Job 0: load holidays ─────────────────────────────────────────────────────
def load_holidays(csv_path: str) -> set:
    df = pd.read_csv(csv_path)
    holidays = set()
    for raw in df['Date'].dropna():
        try:
            holidays.add(datetime.datetime.strptime(str(raw).strip(), '%d.%m.%Y').date())
        except ValueError:
            pass
    print(f"  Loaded {len(holidays)} public holidays.")
    return holidays


# ── Job 1: extend calendar rows ───────────────────────────────────────────────
def extend_calendar(ws, holidays: set, end_dt: datetime.datetime):
    """Append 5-min calendar rows (cols 1-5) up to end_dt."""
    # Find last valid timestamp row
    last_ts = None
    last_row = FIRST_DATA_ROW - 1
    for r in range(ws.max_row, FIRST_DATA_ROW - 1, -1):
        ts = _norm_ts(ws.cell(row=r, column=1).value)
        if ts:
            last_ts = datetime.datetime.strptime(ts, '%d.%m.%Y %H:%M:%S')
            last_row = r
            break

    if last_ts is None:
        print("  WARNING: no existing timestamp found — skipping calendar extension.")
        return

    print(f"  Last existing row : {last_ts.strftime('%d.%m.%Y %H:%M:%S')}  (Excel row {last_row})")

    if last_ts >= end_dt:
        print("  Calendar already covers the requested range — nothing added.")
        return

    next_row = last_row + 1
    current  = last_ts + datetime.timedelta(minutes=STEP_MINUTES)
    count    = 0
    while current <= end_dt:
        d = current.date()
        ws.cell(row=next_row, column=1).value = current.strftime('%d.%m.%Y %H:%M:%S')
        ws.cell(row=next_row, column=2).value = datetime.datetime(d.year, d.month, d.day)
        ws.cell(row=next_row, column=3).value = current.time()
        ws.cell(row=next_row, column=4).value = d.isoweekday()         # Mon=1 … Sun=7
        ws.cell(row=next_row, column=5).value = 1 if d in holidays else 0
        next_row += 1
        current += datetime.timedelta(minutes=STEP_MINUTES)
        count   += 1

    print(f"  Added {count} new calendar rows  "
          f"(up to {end_dt.strftime('%d.%m.%Y %H:%M:%S')}).")


# ── Job 2: build timestamp → Excel row lookup ─────────────────────────────────
def build_ts_index(ws) -> dict:
    """Return {timestamp_string: excel_row_number} for every data row."""
    idx = {}
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        ts = _norm_ts(ws.cell(row=r, column=1).value)
        if ts:
            idx[ts] = r
    return idx


# ── Job 2: build signal → Excel column lookup from row SIGNALNAME_ROW ────────
def build_signal_index(ws) -> dict:
    """Return {signal_name: excel_col_number} from the Signalname row."""
    idx = {}
    for cell in ws[SIGNALNAME_ROW]:
        if cell.value and cell.column > 1:   # skip col 1 ('Signalname' label)
            idx[str(cell.value).strip()] = cell.column
    return idx


# ── Job 2: import one CSV ─────────────────────────────────────────────────────
def import_csv(csv_path: str, ws, ts_index: dict, signal_index: dict):
    """Read one SCADA CSV and write values into the matching Excel rows."""
    fname = os.path.basename(csv_path)
    print(f"\n  Importing: {fname}")

    # Read raw lines — semicolon delimiter, no header inference
    with open(csv_path, encoding='utf-8-sig', errors='replace') as f:
        lines = f.readlines()

    if len(lines) < 3:
        print(f"    SKIP — fewer than 3 lines.")
        return

    # Row 1 (index 0) = units  → skip
    # Row 2 (index 1) = signal names
    signal_row = lines[1].rstrip(';\n').split(';')
    # signal_row[0] = 'Signalname', signal_row[1..] = actual signal names
    csv_signals = [s.strip() for s in signal_row[1:]]

    # Map CSV column index (0-based within data values) → Excel column number
    col_map = {}
    for i, sig in enumerate(csv_signals):
        canonical = SIGNAL_ALIASES.get(sig, sig)
        if canonical in signal_index:
            if canonical != sig:
                print(f"    INFO: alias resolved '{sig}' → '{canonical}'")
            col_map[i] = signal_index[canonical]
        else:
            print(f"    WARNING: signal not found in Excel — '{sig}'")

    if not col_map:
        print(f"    SKIP — no matching signals found.")
        return

    # Parse all data rows into a list of (datetime, {col_idx: float})
    raw_rows = []
    for line in lines[2:]:
        line = line.rstrip(';\n')
        if not line:
            continue
        parts = line.split(';')
        if len(parts) < 2:
            continue
        try:
            ts_dt = datetime.datetime.strptime(parts[0].strip(), '%d.%m.%Y %H:%M:%S')
        except ValueError:
            continue
        vals = {}
        for i, excel_col in col_map.items():
            raw_val = parts[i + 1].strip() if (i + 1) < len(parts) else ''
            if raw_val == '':
                continue
            try:
                vals[i] = float(raw_val.replace(',', '.'))
            except ValueError:
                vals[i] = raw_val
        raw_rows.append((ts_dt, vals))

    if not raw_rows:
        print(f"    SKIP — no valid data rows found.")
        return

    # Detect resolution: gap between first two timestamps
    if len(raw_rows) >= 2:
        gap_minutes = int((raw_rows[1][0] - raw_rows[0][0]).total_seconds() / 60)
    else:
        gap_minutes = STEP_MINUTES

    # Interpolate to 5-min if resolution is coarser (e.g. 15-min)
    if gap_minutes > STEP_MINUTES:
        print(f"    Detected {gap_minutes}-min resolution — interpolating to {STEP_MINUTES}-min ...")
        interp_rows = []
        for k in range(len(raw_rows) - 1):
            t0, v0 = raw_rows[k]
            _,  v1 = raw_rows[k + 1]
            steps = gap_minutes // STEP_MINUTES
            for s in range(steps):
                frac = s / steps
                ts_interp = t0 + datetime.timedelta(minutes=s * STEP_MINUTES)
                vals_interp = {}
                for i, excel_col in col_map.items():
                    a = v0.get(i)
                    b = v1.get(i)
                    # Cols O-R (15-18) are status flags — forward-fill, no interpolation
                    if excel_col in (15, 16, 17, 18):
                        if a is not None:
                            vals_interp[i] = a
                    elif isinstance(a, float) and isinstance(b, float):
                        vals_interp[i] = round(a + frac * (b - a), 3)
                    elif a is not None:
                        vals_interp[i] = a
                interp_rows.append((ts_interp, vals_interp))
        # Append the last original row
        interp_rows.append(raw_rows[-1])
        raw_rows = interp_rows

    # Write to Excel
    written = 0
    skipped = 0
    for ts_dt, vals in raw_rows:
        ts_key    = ts_dt.strftime('%d.%m.%Y %H:%M:%S')
        excel_row = ts_index.get(ts_key)
        if excel_row is None:
            skipped += 1
            continue
        for i, excel_col in col_map.items():
            v = vals.get(i)
            if v is not None:
                ws.cell(row=excel_row, column=excel_col).value = v
        written += 1

    print(f"    Written: {written} rows  |  Skipped (no match): {skipped} rows")


# ── Job 3: data gap check ─────────────────────────────────────────────────────
def check_data_gaps(ws, signal_index: dict, days: int = GAP_CHECK_DAYS):
    """Warn about any day in the last `days` days that has no Load_Is values."""
    load_is_col = signal_index.get(LOAD_IS_SIGNAL)
    if load_is_col is None:
        print("  [Gap check] Load_Is column not found in Excel — skipping.")
        return

    today      = datetime.date.today()
    check_from = today - datetime.timedelta(days=days)
    check_to   = today - datetime.timedelta(days=1)   # yesterday; today not expected yet

    dates_with_data = set()
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        ts = _norm_ts(ws.cell(row=r, column=1).value)
        if ts is None:
            continue
        try:
            d = datetime.datetime.strptime(ts, '%d.%m.%Y %H:%M:%S').date()
        except ValueError:
            continue
        if check_from <= d <= check_to:
            if ws.cell(row=r, column=load_is_col).value is not None:
                dates_with_data.add(d)

    missing = []
    d = check_from
    while d <= check_to:
        if d not in dates_with_data:
            missing.append(d)
        d += datetime.timedelta(days=1)

    if missing:
        print()
        print('!' * 60)
        print('  ATTENTION — missing load data detected in the last 2 weeks:')
        for m in missing:
            print(f"  !  {m.strftime('%A %d.%m.%Y')} — no data found.")
        print('  Please import these days to have a more accurate forecast.')
        print('!' * 60)
        print()
    else:
        print(f"  Data gap check OK — all {days} days present.")


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    today    = datetime.date.today()
    end_date = today + datetime.timedelta(days=EXTEND_DAYS_AHEAD)
    end_dt   = datetime.datetime.combine(end_date, datetime.time(23, 55))

    print("=" * 60)
    print(f"get_load_data  —  {today}")
    print("=" * 60)

    # ── Load holidays ──────────────────────────────────────────────────────
    holidays = load_holidays(HOLIDAY_CSV)

    # ── Open workbook ──────────────────────────────────────────────────────
    print(f"\nOpening: {EXCEL_PATH}")
    wb = load_workbook(EXCEL_PATH)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}")
    ws = wb[SHEET_NAME]

    # ── Job 1: extend calendar ─────────────────────────────────────────────
    print("\n[1] Extending calendar rows ...")
    extend_calendar(ws, holidays, end_dt)

    # ── Job 2: import CSV files ────────────────────────────────────────────
    # signal_index is always built — needed by both the CSV import and gap check
    signal_index = build_signal_index(ws)

    csv_files = sorted(glob.glob(os.path.join(INPUT_DIR, '*.csv')))
    if not csv_files:
        print(f"\n[2] No CSV files found in {INPUT_DIR}")
    else:
        print(f"\n[2] Importing {len(csv_files)} CSV file(s) from Input/ ...")
        ts_index = build_ts_index(ws)
        print(f"    Excel rows indexed : {len(ts_index)}")
        print(f"    Signals in Excel   : {len(signal_index)}")

        for csv_path in csv_files:
            import_csv(csv_path, ws, ts_index, signal_index)

    # ── Job 3: data gap check ──────────────────────────────────────────────
    print("\n[3] Checking data completeness (last 2 weeks) ...")
    check_data_gaps(ws, signal_index)

    # ── Save ───────────────────────────────────────────────────────────────
    wb.save(EXCEL_PATH)
    print(f"\nSaved: {EXCEL_PATH}")


if __name__ == '__main__':
    main()
