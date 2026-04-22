"""
reformat_temp.py
----------------
Converts Meiringen.csv (long format, one row per timestamp+parameter) into
a wide-format .xlsx that matches the structure of Data_Prediction.xlsx:

    Row 1 : Human-readable column names  (Date, Hour, Diffuse Radiation, ...)
    Row 2 : Units                        (-, -, W, W, %, W, mm, %, cm, °C, idx, m/s)
    Row 3 : Original signal / parameter  (-, -, diffuse_rad_mean_15min:W, ...)
    Row 4+: Data rows (one per timestamp after pivot)
"""

import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Paths (always resolved relative to this script's location) ─────────────
HERE       = Path(__file__).parent
INPUT_CSV  = HERE / "Meiringen.csv"
OUTPUT_XLS = HERE / "Meiringen_formatted.xlsx"

# ── Parameter metadata (signal → display name, unit) ──────────────────────
PARAM_META = {
    "diffuse_rad_mean_15min:W":  ("Diffuse Radiation",       "W"),
    "direct_rad_mean_15min:W":   ("Direct Radiation",        "W"),
    "effective_cloud_cover:p":   ("Cloud Cover",             "%"),
    "global_rad_mean_15min:W":   ("Global Radiation",        "W"),
    "precip_15min:mm":           ("Precipitation",           "mm"),
    "relative_humidity_2m:p":    ("Relative Humidity",       "%"),
    "snow_depth:cm":             ("Snow Depth",              "cm"),
    "t_2m:C":                    ("Temperature",             "°C"),
    "total_aod_550nm:idx":       ("Aerosol Optical Depth",   "idx"),
    "wind_speed_10m:ms":         ("Wind Speed",              "m/s"),
}

# ── 1. Load CSV (only useful columns) ─────────────────────────────────────
print("Loading CSV …")
df = pd.read_csv(
    INPUT_CSV,
    usecols=["parameter", "datetime_utc", "value"],
    parse_dates=["datetime_utc"],
)

# ── 2. Pivot long → wide ───────────────────────────────────────────────────
print(f"  {len(df):,} rows | {df['parameter'].nunique()} parameters")
print("Pivoting …")

# Average duplicates (same timestamp + parameter) that may exist
df_wide = (
    df.groupby(["datetime_utc", "parameter"], as_index=False)["value"]
    .mean()
    .pivot(index="datetime_utc", columns="parameter", values="value")
)
df_wide.index.name = None
df_wide = df_wide.sort_index()

# Keep only parameters we have metadata for; preserve their defined order
ordered_params = [p for p in PARAM_META if p in df_wide.columns]
df_wide = df_wide[ordered_params]

print(f"  Pivot shape: {df_wide.shape[0]:,} rows × {df_wide.shape[1]} parameters")

# ── 3. Build header rows ───────────────────────────────────────────────────
display_names = ["Date", "Hour"] + [PARAM_META[p][0] for p in ordered_params]
units         = ["-",    "-"]    + [PARAM_META[p][1] for p in ordered_params]
signals       = ["-",    "-"]    + list(ordered_params)

# ── 4. Write to xlsx ───────────────────────────────────────────────────────
print("Writing xlsx …")
wb = Workbook(write_only=True)
ws = wb.create_sheet("Meiringen")

header_font  = Font(bold=True)
header_fill  = PatternFill("solid", fgColor="D9E1F2")   # light blue, like Data_Prediction
center_align = Alignment(horizontal="center")

def make_header_row(values, bold=False, fill=None):
    from openpyxl.cell import WriteOnlyCell
    cells = []
    for v in values:
        c = WriteOnlyCell(ws, value=v)
        if bold:
            c.font = header_font
        if fill:
            c.fill = fill
        c.alignment = center_align
        cells.append(c)
    return cells

ws.append(make_header_row(display_names, bold=True, fill=header_fill))
ws.append(make_header_row(units,         bold=False, fill=header_fill))
ws.append(make_header_row(signals,       bold=False, fill=header_fill))

# Data rows
for ts, row in df_wide.iterrows():
    date_str = ts.strftime("%d.%m.%Y")
    hour_str = ts.strftime("%H:%M")
    ws.append([date_str, hour_str] + [round(v, 4) if pd.notna(v) else None for v in row])

wb.save(OUTPUT_XLS)
print(f"Done → {OUTPUT_XLS}")
